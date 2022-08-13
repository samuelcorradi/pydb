from storagy.conn import Conn as Super
from storagy.conn.directory import Conn as DirectoryConn
from abc import ABC
import xlrd
from xlrd.sheet import ctype_text
import openpyxl
import re
from os.path import join
import datetime
import string

class Conn(Super):

    class ExcelEngine(ABC):

        def letter_to_numbers(self, col:str, row:str=''):
            col += row
            matches = re.findall(pattern=r'([A-Za-z]+)([0-9]+)', string=col)
            if matches:
                return matches[0]

    class XLSEngine(ExcelEngine):

        def __init__(self, filepath:str, sheet=0):
            # start important for Python 3.9+
            xlrd.xlsx.ensure_elementtree_imported(False, None)
            xlrd.xlsx.Element_has_iter = True
            # end important for Python 3.9+
            self._filepath = filepath
            self._wb = xlrd.open_workbook(self._filepath)
            self._sheet = self._wb.sheet_by_name(sheet)
        
        def open(self):
            pass
        
        def sheetnames(self):
            """
            Pega a lista de planilhas em uma
            pasta de trabalho.
            """
            return self._wb.sheet_names()

        def nrows(self):
            """
            Numero de linhas em uma planilha.
            """
            return self._sheet.nrows

        def read_line(self, line=0):
            """
            Retorna o conteu da linha indicada.
            """
            return self._sheet.row_values(line)
        
        def read(self, begin=None, end=None):
            """
            Le um intervalo.
            """
            pass

        def write_cell(self, val, line:int=0, col:int=0):
            pass

        def sheet_names(self):
            return self._wb.sheet_names()

    class XLSXEngine(ExcelEngine):
        
        def __init__(self, filepath:str, sheet=0):
            self._filepath = filepath
            self._wb = openpyxl.load_workbook(filename=self._filepath)
            self._sheet = sheet

        def sheetnames(self):
            """
            Pega a lista de planilhas em uma
            pasta de trabalho.
            """
            return self._wb.sheetnames

        def read_line(self, line=0):
            """
            Retorna o conteu da linha indicada.
            """
            max_col = self._wb[self._sheet].max_column
            max_row = self._wb[self._sheet].max_row
            #row = []
            # for i in range(1, max_row+1):
            cols = []
            for i in range(1, max_col+1):
                cell_obj = self._wb[self._sheet].cell(row=line+1, column=i)
                cols.append(cell_obj.value)
                # row.append(cols)
            return cols

        def nrows(self):
            """
            Numero de linhas em uma planilha.
            """
            return self._wb[self._sheet].max_row

        def read(self, sheet=0, begin=None, end=None):
            """
            Le um intervalo.
            """
            sheet_ranges = self._wb[sheet]
            _end = ':' + ''.join(end) if end else ''
            interval = ''.join(begin) + _end
            data = sheet_ranges[interval].value
            return data

        def read_cell(self, row:int=1, col:int=1):
            cell_obj = self._wb[self._sheet].cell(row=row, column=col)
            return cell_obj.value

        def write_cell(self, value, row:int=1, col:int=1):
            self._wb[self._sheet].cell(column=col, row=row, value=value)

        def sheet_names(self):
            return self._wb.sheetnames
        

    @staticmethod
    def sources(path:str
        , filename:str
        , filter:str=None)->list:
        conn = Conn(path=path
            , filename=filename
            , sheet=0)
        sheet_names = conn.sheet_names()
        r = [f for f in sheet_names if not filter or f.startswith(filter)]
        conn = None
        return r

    def __init__(self
        , path:str
        , filename:str
        , sheet
        , range=None
        , has_header:bool=True
        , encode:str='utf-8'):
        """
        sheet can be a name or int position.
        """
        dir_conn = DirectoryConn(path)
        self._filename = filename
        self._filepath = dir_conn._path.append_file(filename) # join(path, filename)
        self._wb = None
        self._sheet = sheet
        self._has_header = has_header
        self._encode = encode
        self.set_range(range)
        super().__init__()

    def _connect(self):
        """
        ...
        """
        return self.load_drive(self._filepath)

    def _disconnect(self):
        """
        ...
        """
        self._handler = None

    @staticmethod
    def col2letter(col:int)->str:
        """
        Receives a number and returns
        the corresponding letter.
        https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
        """
        if col<1:
            raise Exception("Collumn number should be 1 or bigger. Value '{}' is not valid.".format(col))
        letters = ''
        while col:
            mod = (col - 1) % 26
            letters += chr(mod + 65)
            col = (col - 1) // 26
        return ''.join(reversed(letters))

    @staticmethod
    def col2number(col:str)->int:
        """
        ...
        """
        if not col or type(col) is not str:
            return 0
        col_pos = 0
        _tb = string.ascii_uppercase
        matches = re.findall(pattern=r'([A-Za-z])', string=col)
        if not matches:
            raise Exception("Digite uma letra.")
        for m in matches:
            col_pos = col_pos * 26 + (ord(m.upper()) - ord('A')) + 1
        return col_pos

    @staticmethod
    def range_from_str(r:str)->dict:
        """
        ...
        """
        matches = re.findall(pattern=r'([A-Za-z]+)([0-9]+)(?:\:([A-Za-z]+)?([0-9]+)?)?', string=r)
        if matches:
            _range = list(matches[0])
            _range[0] = Conn.col2number(_range[0])
            _range[1] = int(_range[1])
            _range[2] = Conn.col2number(_range[2])
            _range[3] = int(_range[3]) if _range[3] else 0
            return tuple(_range)
        return (0, 0, 0, 0)

    @staticmethod
    def range_from_tuple(r:tuple)->dict:
        """
        ...
        """
        _range = []
        defautl = (0, 0, 0, 0)
        for i in list(range(4)):
            _range.append(r[i] if i<len(r) else defautl[i])
        _range[0] = _range[0] if type(_range[0]) is int else Conn.col2number(_range[0])
        _range[2] = _range[2] if type(_range[2]) is int else Conn.col2number(_range[2])
        return tuple(_range)

    def sheet_names(self):
        return self.get_handler().sheet_names()

    def get_range_size(self)->tuple:
        _range = self.get_range()
        return abs(_range[3]-_range[1])+1 if _range[3] else 0\
            , abs(_range[2]-_range[0])+1 if _range[2] else 0

    def get_range(self):
        """
        Getter method to access sheet's range.
        """
        return self._range

    def set_range(self, range):
        """
        Indicates the ranges that define
        the area where the data is on
        the worksheet.
        """
        _range = (0, 0, 0, 0)
        if type(range) is str:
            _range = Conn.range_from_str(range)
        if type(range) is tuple:
            _range = Conn.range_from_tuple(range)
        self._range = _range

    def _check_extension(self, filename:str):
        """
        ...
        """
        if '.xlsx' in filename:
            return 'xlsx'
        elif '.xls':
            return 'xls'

    def load_drive(self, filepath):
        """
        ...
        """
        ext = self._check_extension(self._filename)
        engine = None
        print(ext)
        if ext=='xls':
            engine = self.XLSEngine(self._filepath, self._sheet)
        elif ext=='xlsx':
            engine = self.XLSXEngine(self._filepath, self._sheet)
        return engine

    def nrows(self)->int:
        """
        ...
        """
        nrows = self._handler.nrows()
        return nrows if not self._has_header else nrows-1

    def get_filepath(self)->str:
        """
        ...
        """
        return self._filepath

    def get_filename(self)->str:
        """
        ...
        """
        return self._filename

    def get_type_list(self):
        """
        ...
        """
        type_list = []
        row_idx = int(self._has_header)
        row = self._handler.row(row_idx)
        for cell_obj in row:
            cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
            type_list.append(cell_type_str)
        return type_list

    def insert(self, data):
        """
        ...
        """
        _range = self.get_range()
        pass

    def field_list(self):
        """
        ...
        """
        _range = self.get_range()
        range_size = self.get_range_size()
        first_row = _range[1] if _range[1] else 1
        # if have header
        if self._has_header:
            # if range is definied
            if range_size[0]:
                r=[]
                #print(list(range(_range[0], _range[2])))
                for i in list(range(_range[0], _range[2]+1)):
                    r.append(self._handler.read_cell(row=first_row, col=i))
                return r
            # if range is NOT definied
            else:
                return self._handler.read_line(0)
        # if have NO header
        else:
            r=[]
            if range_size[0]:
                for i in list(range(_range[0], _range[2]+1)):
                    r.append(Conn.col2letter(i))
            else:
                line = self._handler.read_line(first_row)
                for i, _ in enumerate(line):
                    r.append(Conn.col2letter(i+1))
            return r

    def rename(self):
        """
        ...
        """
        pass

    def is_empty(self):
        """
        ...
        """
        pass

    def truncate(self):
        """
        ...
        """
        pass

    def erase(self):
        """
        ...
        """
        pass

    def select(self):
        """
        ...
        """
        pass

    def get_col_names(self):
        """
        Why??
        """
        ncols = len(self.field_list()) # self._handler.ncols
        if self._has_header:
            cols = self._handler.read_line(0)
            if len(cols)<ncols:
                cols += [str(i) for i in range(len(cols), ncols)]
            return cols
        return [str(i) for i in range(1, ncols+1)]

    def __parse_row_value(self, row:list, type_list:list):
        return [self.__parse_cell_value(j, v, type_list) for j, v in enumerate(row)]

    def __parse_cell_value(self, col_idx:int, value, type_list:str):
        type = type_list[col_idx]
        if type=='xldate':
            return datetime.datetime(*xlrd.xldate_as_tuple(value, self._wb.datemode))
        elif type=='empty':
            return None
        elif type=='number':
            return float(value)
        return value

    def _get_row(self, row_num:int, type_list:list):
        row = self._handler.row_values(row_num)
        return [self.__parse_cell_value(col_idx, v, type_list) for col_idx, v in enumerate(row)]

    def all(self)->list:
        """
        Metodo para carregar o dataset.
        """
        r = []
        # type_list = self.get_type_list()
        first_row = 1 if self._has_header else 0
        for row_num in range(first_row, self._handler.nrows()):
            r.append(self._handler.read_line(row_num))
            # r.append(self._get_row(row_num, type_list))
        return r